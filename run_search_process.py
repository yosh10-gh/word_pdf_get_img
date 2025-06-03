import os
import glob
import pandas as pd
import zipfile
from PIL import Image
import io
import tempfile
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_docx_pdf_files(directory_path):
    """
    1. ディレクトリを全てクロールし、.docx, .pdf ファイルを抽出
    
    Args:
        directory_path (str): 検索対象のディレクトリパス
        
    Returns:
        list: 見つかった .docx と .pdf ファイルのパスのリスト
    """
    found_files = []
    
    try:
        print(f"🔍 ステップ1: ディレクトリクロール開始")
        print(f"対象ディレクトリ: {os.path.abspath(directory_path)}")
        print("-" * 60)
        
        # ディレクトリの存在確認
        if not os.path.exists(directory_path):
            print(f"❌ エラー: ディレクトリが存在しません: {directory_path}")
            return found_files
        
        # .docx ファイルを検索
        docx_pattern = os.path.join(directory_path, "**", "*.docx")
        docx_files = glob.glob(docx_pattern, recursive=True)
        
        # .pdf ファイルを検索
        pdf_pattern = os.path.join(directory_path, "**", "*.pdf")
        pdf_files = glob.glob(pdf_pattern, recursive=True)
        
        # 有効なファイルのみを追加
        for file_path in docx_files + pdf_files:
            if os.path.isfile(file_path) and os.access(file_path, os.R_OK):
                found_files.append(file_path)
                print(f"  📄 発見: {file_path}")
        
        print("-" * 60)
        print(f"✅ ステップ1完了: 合計 {len(found_files)} ファイル発見 (DOCX: {len(docx_files)}, PDF: {len(pdf_files)})")
        
    except Exception as e:
        print(f"❌ ファイル検索エラー: {str(e)}")
    
    return found_files

def has_images_in_docx(file_path):
    """
    .docxファイルに画像が含まれているかチェックする
    """
    try:
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return False
            
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            for file_info in zip_file.filelist:
                if file_info.filename.startswith('word/media/'):
                    return True
        return False
    except (zipfile.BadZipFile, PermissionError, OSError):
        return False
    except Exception:
        return False

def has_images_in_pdf(file_path):
    """
    .pdfファイルに画像が含まれているかチェックする
    """
    try:
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return False
            
        import PyPDF2
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            if pdf_reader.is_encrypted:
                try:
                    pdf_reader.decrypt('')
                except:
                    return False
            
            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    if '/Resources' in page and '/XObject' in page['/Resources']:
                        xObject = page['/Resources']['/XObject'].get_object()
                        for obj in xObject:
                            if xObject[obj]['/Subtype'] == '/Image':
                                return True
                except Exception:
                    continue
        return False
    except Exception:
        return False

def filter_files_with_images(file_list):
    """
    2. 抽出したものから、imgファイルが含まれているものだけをさらに抽出
    
    Args:
        file_list (list): ファイルパスのリスト
        
    Returns:
        list: 画像が含まれているファイルのリスト
    """
    files_with_images = []
    
    print(f"\n🖼️ ステップ2: 画像含有ファイル判定開始")
    print("-" * 60)
    
    for i, file_path in enumerate(file_list, 1):
        try:
            print(f"  {i:2d}. チェック中: {os.path.basename(file_path)}")
            
            if file_path.endswith('.docx'):
                if has_images_in_docx(file_path):
                    files_with_images.append(file_path)
                    print(f"      ✅ 画像あり (DOCX)")
                else:
                    print(f"      ❌ 画像なし (DOCX)")
                    
            elif file_path.endswith('.pdf'):
                if has_images_in_pdf(file_path):
                    files_with_images.append(file_path)
                    print(f"      ✅ 画像あり (PDF)")
                else:
                    print(f"      ❌ 画像なし (PDF)")
        except Exception as e:
            print(f"      ⚠️ ファイルチェックエラー: {str(e)}")
            continue
    
    print("-" * 60)
    print(f"✅ ステップ2完了: {len(files_with_images)}/{len(file_list)} ファイルに画像が含まれています")
    
    return files_with_images

def get_media_filenames(file_path):
    """
    .docxファイル内の画像ファイル名のリストを取得する
    """
    media_filenames = []
    try:
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return media_filenames
            
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            for file_info in zip_file.filelist:
                if file_info.filename.startswith('word/media/') and not file_info.is_dir():
                    filename = os.path.basename(file_info.filename)
                    media_filenames.append(filename)
        
        media_filenames.sort()
        
    except Exception:
        pass
    return media_filenames

def extract_images_from_docx(file_path):
    """
    3. .docxは内部データを展開して画像を抽出
    """
    images = []
    try:
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return images
            
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            for file_info in zip_file.filelist:
                if file_info.filename.startswith('word/media/') and not file_info.is_dir():
                    try:
                        image_data = zip_file.read(file_info.filename)
                        if len(image_data) == 0:
                            continue
                            
                        image = Image.open(io.BytesIO(image_data))
                        if image.mode in ('RGBA', 'LA', 'P'):
                            image = image.convert('RGB')
                        images.append(image)
                    except (Image.UnidentifiedImageError, OSError):
                        continue
                    except Exception:
                        continue
    except Exception:
        pass
    return images

def extract_images_from_pdf(file_path):
    """
    PDFファイルから画像を抽出する
    """
    images = []
    try:
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return images
            
        import PyPDF2
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            if pdf_reader.is_encrypted:
                try:
                    pdf_reader.decrypt('')
                except:
                    return images
            
            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    if '/Resources' in page and '/XObject' in page['/Resources']:
                        xObject = page['/Resources']['/XObject'].get_object()
                        for obj in xObject:
                            if xObject[obj]['/Subtype'] == '/Image':
                                try:
                                    img_obj = xObject[obj]
                                    if '/Filter' in img_obj:
                                        if img_obj['/Filter'] == '/DCTDecode':
                                            img_data = img_obj._data
                                            if len(img_data) > 0:
                                                image = Image.open(io.BytesIO(img_data))
                                                if image.mode in ('RGBA', 'LA', 'P'):
                                                    image = image.convert('RGB')
                                                images.append(image)
                                        elif img_obj['/Filter'] == '/FlateDecode':
                                            try:
                                                width = int(img_obj['/Width'])
                                                height = int(img_obj['/Height'])
                                                img_data = img_obj._data
                                                if len(img_data) > 0 and '/ColorSpace' in img_obj:
                                                    if img_obj['/ColorSpace'] == '/DeviceRGB':
                                                        expected_size = width * height * 3
                                                        if len(img_data) >= expected_size:
                                                            image = Image.frombytes('RGB', (width, height), img_data[:expected_size])
                                                            images.append(image)
                                            except (ValueError, TypeError):
                                                continue
                                except Exception:
                                    continue
                except Exception:
                    continue
    except Exception:
        pass
    return images

def resize_image_to_100px(image):
    """
    5. 画像を100px×100pxにリサイズする（アスペクト比を保持）
    """
    try:
        if image.size[0] == 0 or image.size[1] == 0:
            raise ValueError("無効な画像サイズ")
        
        image.thumbnail((100, 100), Image.Resampling.LANCZOS)
        return image
    except Exception:
        fallback_image = Image.new('RGB', (50, 50), 'white')
        return fallback_image

def create_excel_with_images(file_list, output_dir="result", output_filename="検索結果.xlsx"):
    """
    4-5. 新規Excelを作成し、ファイルパスをA列に並べ、
         埋め込まれているimgファイルを100px×100pxで表示してresultディレクトリに出力
    
    Args:
        file_list (list): 画像含有ファイルのリスト
        output_dir (str): 出力ディレクトリ
        output_filename (str): 出力するExcelファイル名
    """
    if not file_list:
        print("❌ 保存するファイルがありません。")
        return
    
    print(f"\n📊 ステップ4-5: Excel作成・画像表示開始")
    print("-" * 60)
    
    # 出力ディレクトリを作成
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, output_filename)
    
    # 新しいワークブックを作成
    wb = Workbook()
    ws = wb.active
    
    # 全ファイルの画像ファイル名を調査
    all_image_filenames = set()
    file_image_map = {}
    
    print("  🔍 画像ファイル名調査中...")
    for file_path in file_list:
        try:
            if file_path.endswith('.docx'):
                media_filenames = get_media_filenames(file_path)
                file_image_map[file_path] = media_filenames
                all_image_filenames.update(media_filenames)
                print(f"    📄 {os.path.basename(file_path)}: {len(media_filenames)}個の画像")
            elif file_path.endswith('.pdf'):
                images = extract_images_from_pdf(file_path)
                pdf_filenames = [f"pdf_image{i+1}" for i in range(len(images))]
                file_image_map[file_path] = pdf_filenames
                all_image_filenames.update(pdf_filenames)
                print(f"    📑 {os.path.basename(file_path)}: {len(images)}個の画像")
        except Exception:
            file_image_map[file_path] = []
    
    # 画像ファイル名をソート
    sorted_image_filenames = sorted(all_image_filenames)
    max_images = len(sorted_image_filenames)
    
    print(f"    📊 ユニークな画像ファイル名数: {max_images}")
    
    # ヘッダーを動的に設定
    ws['A1'] = 'ファイルパス'
    for i, filename in enumerate(sorted_image_filenames):
        col_letter = chr(ord('B') + i) if i < 25 else f"A{chr(ord('A') + i - 25)}"
        ws[f'{col_letter}1'] = filename
    
    # 行の高さを設定（100px用）
    for row in range(2, len(file_list) + 2):
        ws.row_dimensions[row].height = 75
    
    # 列の幅を設定
    ws.column_dimensions['A'].width = 50
    for i in range(max_images):
        col_letter = chr(ord('B') + i) if i < 25 else f"A{chr(ord('A') + i - 25)}"
        ws.column_dimensions[col_letter].width = 15
    
    temp_files = []
    
    try:
        print("  🖼️ 画像抽出・配置中...")
        for idx, file_path in enumerate(file_list, start=2):
            try:
                print(f"    {idx-1:2d}. 処理中: {os.path.basename(file_path)}")
                
                # A列にファイルパスを設定
                ws[f'A{idx}'] = file_path
                
                # 画像を抽出
                images = []
                current_filenames = file_image_map.get(file_path, [])
                
                if file_path.endswith('.docx'):
                    images = extract_images_from_docx(file_path)
                elif file_path.endswith('.pdf'):
                    images = extract_images_from_pdf(file_path)
                
                # 実際のファイル名と画像を対応付けて配置
                for img_idx, (filename, image) in enumerate(zip(current_filenames, images)):
                    try:
                        if filename in sorted_image_filenames:
                            col_idx = sorted_image_filenames.index(filename)
                            col_letter = chr(ord('B') + col_idx) if col_idx < 25 else f"A{chr(ord('A') + col_idx - 25)}"
                            
                            # 画像を100px×100pxにリサイズ
                            resized_image = resize_image_to_100px(image.copy())
                            
                            # 一時ファイルを作成
                            temp_fd, temp_path = tempfile.mkstemp(suffix='.png')
                            temp_files.append(temp_path)
                            
                            try:
                                os.close(temp_fd)
                                resized_image.save(temp_path, 'PNG', optimize=True)
                                
                                if os.path.getsize(temp_path) > 10 * 1024 * 1024:
                                    continue
                                
                                # Excelに画像を挿入
                                img = OpenpyxlImage(temp_path)
                                img.width = 100
                                img.height = 100
                                
                                cell_location = f'{col_letter}{idx}'
                                ws.add_image(img, cell_location)
                                
                            except Exception:
                                pass
                    except Exception:
                        continue
                
                print(f"        ✅ {len(images)}個の画像を配置完了")
            except Exception:
                continue
        
        # Excelファイルを保存
        try:
            wb.save(output_path)
            print(f"\n✅ ステップ4-5完了: Excelファイルを保存しました")
            print(f"   📁 出力パス: {os.path.abspath(output_path)}")
        except (PermissionError, OSError) as e:
            import time
            alt_filename = f"検索結果_{int(time.time())}.xlsx"
            alt_path = os.path.join(output_dir, alt_filename)
            wb.save(alt_path)
            print(f"✅ 代替ファイル名で保存: {os.path.abspath(alt_path)}")
    
    finally:
        # 一時ファイルをクリーンアップ
        for temp_path in temp_files:
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception:
                pass

def main():
    """
    test_directoryを対象として、1番から検索結果完了まで実行
    """
    # 検索対象のディレクトリを指定
    search_directory = "test_directory"  # ユーザー指定（修正）
    
    print("🚀 検索結果出力プロセス開始")
    print("=" * 80)
    
    # 1. ディレクトリクロール
    files = extract_docx_pdf_files(search_directory)
    
    if not files:
        print("❌ 該当するファイルが見つかりませんでした。")
        return
    
    # 2. 画像含有ファイル判定
    files_with_images = filter_files_with_images(files)
    
    if not files_with_images:
        print("❌ 画像が含まれているファイルがありませんでした。")
        return
    
    # 3-5. DOCX内部データ展開 + Excel作成 + 画像表示・出力
    create_excel_with_images(files_with_images, "result", "検索結果.xlsx")
    
    print("\n" + "=" * 80)
    print("🎉 全プロセス完了！")
    print("=" * 80)
    print(f"📊 最終結果:")
    print(f"   - 検索対象: {search_directory}")
    print(f"   - 発見ファイル数: {len(files)}")
    print(f"   - 画像含有ファイル数: {len(files_with_images)}")
    print(f"   - 出力ファイル: result/検索結果.xlsx")
    print("=" * 80)

if __name__ == "__main__":
    main() 