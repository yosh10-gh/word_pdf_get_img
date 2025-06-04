import os
import glob
import pandas as pd
import zipfile
from PIL import Image
import io
import tempfile
import shutil
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_docx_pdf_files(directory_path):
    """
    指定されたディレクトリを再帰的にクロールして、
    .docx と .pdf ファイルのパスを抽出する
    
    Args:
        directory_path (str): 検索対象のディレクトリパス
        
    Returns:
        list: 見つかった .docx と .pdf ファイルのパスのリスト
    """
    found_files = []
    
    try:
        # ディレクトリの存在確認
        if not os.path.exists(directory_path):
            print(f"警告: ディレクトリが存在しません: {directory_path}")
            return found_files
        
        # .docx ファイルを検索
        docx_pattern = os.path.join(directory_path, "**", "*.docx")
        docx_files = glob.glob(docx_pattern, recursive=True)
        # 有効なファイルのみを追加
        for file_path in docx_files:
            if os.path.isfile(file_path) and os.access(file_path, os.R_OK):
                found_files.append(file_path)
        
        # .pdf ファイルを検索
        pdf_pattern = os.path.join(directory_path, "**", "*.pdf")
        pdf_files = glob.glob(pdf_pattern, recursive=True)
        # 有効なファイルのみを追加
        for file_path in pdf_files:
            if os.path.isfile(file_path) and os.access(file_path, os.R_OK):
                found_files.append(file_path)
    
    except Exception as e:
        print(f"ファイル検索エラー: {str(e)}")
    
    return found_files

def has_images_in_docx(file_path):
    """
    .docxファイルに画像が含まれているかチェックする
    
    Args:
        file_path (str): .docxファイルのパス
        
    Returns:
        bool: 画像が含まれている場合True
    """
    try:
        # ファイルの存在とアクセス権限をチェック
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return False
            
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            # word/media/ フォルダ内のファイルをチェック
            for file_info in zip_file.filelist:
                if file_info.filename.startswith('word/media/'):
                    return True
        return False
    except (zipfile.BadZipFile, PermissionError, OSError) as e:
        print(f"  → docxファイル読み込みエラー: {str(e)}")
        return False
    except Exception:
        return False

def get_media_filenames(file_path):
    """
    .docxファイル内の画像ファイル名のリストを取得する（実際の順序を保持）
    
    Args:
        file_path (str): .docxファイルのパス
        
    Returns:
        list: 画像ファイル名のリスト（Word内の実際の順序）
    """
    media_filenames = []
    try:
        # ファイルの存在とアクセス権限をチェック
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return media_filenames
            
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            # word/media/ フォルダ内のファイルを実際の順序で取得
            for file_info in zip_file.filelist:
                if file_info.filename.startswith('word/media/') and not file_info.is_dir():
                    # ファイル名のみを取得
                    filename = os.path.basename(file_info.filename)
                    media_filenames.append(filename)
        
    except (zipfile.BadZipFile, PermissionError, OSError) as e:
        print(f"  → docxファイル処理エラー: {str(e)}")
    except Exception:
        pass
    return media_filenames

def extract_images_from_docx(file_path):
    """
    .docxファイルから画像を抽出する（Word内の実際の順序を保持）
    
    Args:
        file_path (str): .docxファイルのパス
        
    Returns:
        list: 抽出された画像のPILImageオブジェクトのリスト（Word内の実際の順序）
    """
    images = []
    try:
        # ファイルの存在とアクセス権限をチェック
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return images
            
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            # Word内の実際の順序でファイルを取得
            media_files = []
            for file_info in zip_file.filelist:
                if file_info.filename.startswith('word/media/') and not file_info.is_dir():
                    media_files.append(file_info.filename)
            
            # 実際の順序で画像を処理
            for filename in media_files:
                try:
                    # 画像データを読み込み
                    image_data = zip_file.read(filename)
                    if len(image_data) == 0:  # 空のファイルをスキップ
                        # プレースホルダー画像を作成
                        placeholder = Image.new('RGB', (100, 100), color='lightgray')
                        images.append(placeholder)
                        continue
                        
                    # PILImageオブジェクトに変換
                    try:
                        image = Image.open(io.BytesIO(image_data))
                        # 画像形式を確認してRGBに変換
                        if image.mode in ('RGBA', 'LA', 'P'):
                            image = image.convert('RGB')
                        images.append(image)
                    except (Image.UnidentifiedImageError, OSError) as e:
                        print(f"    → 画像読み込みエラー ({filename}): {str(e)}")
                        # エラーの場合はプレースホルダー画像を作成
                        placeholder = Image.new('RGB', (100, 100), color='lightgray')
                        # エラーメッセージをテキストとして追加
                        from PIL import ImageDraw, ImageFont
                        draw = ImageDraw.Draw(placeholder)
                        try:
                            # フォントサイズを小さく設定
                            draw.text((10, 40), "読込不可", fill='black')
                            draw.text((10, 55), os.path.basename(filename)[:10], fill='black')
                        except:
                            pass
                        images.append(placeholder)
                        continue
                except Exception as e:
                    print(f"    → ファイル処理エラー ({filename}): {str(e)}")
                    # エラーの場合はプレースホルダー画像を作成
                    placeholder = Image.new('RGB', (100, 100), color='lightgray')
                    images.append(placeholder)
                    continue
    except (zipfile.BadZipFile, PermissionError, OSError) as e:
        print(f"  → docxファイル処理エラー: {str(e)}")
    except Exception:
        pass
    return images

def has_images_in_pdf(file_path):
    """
    .pdfファイルに画像が含まれているかチェックする
    
    Args:
        file_path (str): .pdfファイルのパス
        
    Returns:
        bool: 画像が含まれている場合True
    """
    try:
        # ファイルの存在とアクセス権限をチェック
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return False
            
        import PyPDF2
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            # 暗号化されたPDFの場合
            if pdf_reader.is_encrypted:
                try:
                    pdf_reader.decrypt('')  # 空のパスワードで試行
                except:
                    print(f"  → 暗号化されたPDFです: {file_path}")
                    return False
            
            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    if '/Resources' in page and '/XObject' in page['/Resources']:
                        xObject = page['/Resources']['/XObject'].get_object()
                        for obj in xObject:
                            if xObject[obj]['/Subtype'] == '/Image':
                                return True
                except Exception:
                    continue
        return False
    except (PyPDF2.errors.PdfReadError, PermissionError, OSError) as e:
        print(f"  → PDFファイル読み込みエラー: {str(e)}")
        return False
    except Exception:
        return False

def extract_images_from_pdf(file_path):
    """
    .pdfファイルから画像を抽出する
    
    Args:
        file_path (str): .pdfファイルのパス
        
    Returns:
        list: 抽出された画像のPILImageオブジェクトのリスト
    """
    images = []
    try:
        # ファイルの存在とアクセス権限をチェック
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return images
            
        import PyPDF2
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            # 暗号化されたPDFの場合
            if pdf_reader.is_encrypted:
                try:
                    pdf_reader.decrypt('')  # 空のパスワードで試行
                except:
                    print(f"  → 暗号化されたPDFをスキップ: {file_path}")
                    return images
            
            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    if '/Resources' in page and '/XObject' in page['/Resources']:
                        xObject = page['/Resources']['/XObject'].get_object()
                        for obj in xObject:
                            if xObject[obj]['/Subtype'] == '/Image':
                                try:
                                    # 画像データを抽出
                                    img_obj = xObject[obj]
                                    if '/Filter' in img_obj:
                                        if img_obj['/Filter'] == '/DCTDecode':
                                            # JPEG画像
                                            img_data = img_obj._data
                                            if len(img_data) > 0:  # 空のデータをスキップ
                                                image = Image.open(io.BytesIO(img_data))
                                                if image.mode in ('RGBA', 'LA', 'P'):
                                                    image = image.convert('RGB')
                                                images.append(image)
                                        elif img_obj['/Filter'] == '/FlateDecode':
                                            # PNG/その他の圧縮画像
                                            try:
                                                width = int(img_obj['/Width'])
                                                height = int(img_obj['/Height'])
                                                img_data = img_obj._data
                                                if len(img_data) > 0 and '/ColorSpace' in img_obj:
                                                    if img_obj['/ColorSpace'] == '/DeviceRGB':
                                                        expected_size = width * height * 3
                                                        if len(img_data) >= expected_size:
                                                            image = Image.frombytes('RGB', (width, height), img_data[:expected_size])
                                                            images.append(image)
                                            except (ValueError, TypeError):
                                                continue
                                except (Image.UnidentifiedImageError, OSError, ValueError) as e:
                                    print(f"    → PDF画像抽出エラー (page {page_num + 1}): {str(e)}")
                                    continue
                                except Exception:
                                    continue
                except Exception:
                    continue
    except (PyPDF2.errors.PdfReadError, PermissionError, OSError) as e:
        print(f"  → PDFファイル処理エラー: {str(e)}")
    except Exception:
        pass
    return images

def extract_docx_structure(file_path, output_base_dir="extracted_structures"):
    """
    .docxファイルの内部構造を指定されたディレクトリに展開する
    
    Args:
        file_path (str): .docxファイルのパス
        output_base_dir (str): 展開先のベースディレクトリ
        
    Returns:
        str: 展開されたディレクトリのパス（失敗時はNone）
    """
    try:
        # ファイルの存在とアクセス権限をチェック
        if not os.path.exists(file_path) or not os.access(file_path, os.R_OK):
            return None
            
        # 出力ディレクトリを作成
        os.makedirs(output_base_dir, exist_ok=True)
        
        # ファイル名から展開先ディレクトリ名を作成
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        # 相対パスの情報も含める（安全な文字に変換）
        relative_path = os.path.relpath(file_path, ".")
        safe_path = relative_path.replace("\\", "_").replace("/", "_").replace(":", "_")
        safe_name = safe_path.replace(".docx", "")
        
        extract_dir = os.path.join(output_base_dir, safe_name)
        
        # 既存のディレクトリが存在する場合は削除
        if os.path.exists(extract_dir):
            shutil.rmtree(extract_dir)
        
        # zipファイルとして展開
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            zip_file.extractall(extract_dir)
        
        print(f"  → 内部構造を展開: {extract_dir}")
        return extract_dir
        
    except (zipfile.BadZipFile, PermissionError, OSError) as e:
        print(f"  → docx展開エラー: {str(e)}")
        return None
    except Exception as e:
        print(f"  → 予期しないエラー: {str(e)}")
        return None

def filter_files_with_images(file_list):
    """
    ファイルリストから画像が含まれているファイルのみを抽出する
    
    Args:
        file_list (list): ファイルパスのリスト
        
    Returns:
        list: 画像が含まれているファイルのリスト
    """
    files_with_images = []
    
    for file_path in file_list:
        try:
            print(f"チェック中: {file_path}")
            
            if file_path.endswith('.docx'):
                if has_images_in_docx(file_path):
                    files_with_images.append(file_path)
                    print(f"  → 画像あり")
                    # Word文書の内部構造を展開
                    extract_docx_structure(file_path)
                else:
                    print(f"  → 画像なし")
                    
            elif file_path.endswith('.pdf'):
                if has_images_in_pdf(file_path):
                    files_with_images.append(file_path)
                    print(f"  → 画像あり")
                else:
                    print(f"  → 画像なし")
        except Exception as e:
            print(f"  → ファイルチェックエラー: {str(e)}")
            continue
    
    return files_with_images

def resize_image_to_100px(image):
    """
    画像を100px×100pxにリサイズする（アスペクト比を保持）
    
    Args:
        image: PILImageオブジェクト
        
    Returns:
        PILImageオブジェクト: リサイズされた画像
    """
    try:
        # 画像のサイズをチェック
        if image.size[0] == 0 or image.size[1] == 0:
            raise ValueError("無効な画像サイズ")
        
        # アスペクト比を保持して100px以内にリサイズ
        image.thumbnail((100, 100), Image.Resampling.LANCZOS)
        return image
    except Exception as e:
        print(f"    → 画像リサイズエラー: {str(e)}")
        # フォールバック: 最小サイズの白い画像を作成
        fallback_image = Image.new('RGB', (50, 50), 'white')
        return fallback_image

def save_to_excel_with_images(file_list, output_dir="result", output_filename="検索結果.xlsx"):
    """
    ファイルリストと画像をExcelファイルに保存する（各ファイルごとに2行使用：上段に画像ファイル名、下段にファイルパスと画像）
    
    Args:
        file_list (list): ファイルパスのリスト
        output_dir (str): 出力ディレクトリ
        output_filename (str): 出力するExcelファイル名
    """
    if not file_list:
        print("保存するファイルがありません。")
        return
    
    # 出力ディレクトリを作成
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, output_filename)
    
    # 新しいワークブックを作成
    wb = Workbook()
    ws = wb.active
    
    # 全ファイルの最大画像数を調査してヘッダーを決定
    max_images = 0
    print("各ファイルの画像数を調査中...")
    for file_path in file_list:
        try:
            if file_path.endswith('.docx'):
                media_filenames = get_media_filenames(file_path)
                max_images = max(max_images, len(media_filenames))
            elif file_path.endswith('.pdf'):
                images = extract_images_from_pdf(file_path)
                max_images = max(max_images, len(images))
        except Exception:
            continue
    
    print(f"最大画像数: {max_images}")
    
    # ヘッダーを設定
    ws['A1'] = 'ファイルパス'
    # B列以降のヘッダーは空白にする（image1, image2などは不要）
    for i in range(max_images):
        col_letter = chr(ord('B') + i) if i < 25 else f"A{chr(ord('A') + i - 25)}"
        ws[f'{col_letter}1'] = ''  # 空白に変更
    
    # 列の幅を設定
    ws.column_dimensions['A'].width = 70
    for i in range(max_images):
        col_letter = chr(ord('B') + i) if i < 25 else f"A{chr(ord('A') + i - 25)}"
        ws.column_dimensions[col_letter].width = 15
    
    temp_files = []  # 一時ファイルのリストを保持
    
    try:
        current_row = 2  # ヘッダーの次の行から開始
        
        for file_path in file_list:
            try:
                print(f"画像抽出中: {file_path}")
                
                # 画像を抽出
                images = []
                image_filenames = []
                
                if file_path.endswith('.docx'):
                    images = extract_images_from_docx(file_path)
                    image_filenames = get_media_filenames(file_path)
                elif file_path.endswith('.pdf'):
                    images = extract_images_from_pdf(file_path)
                    image_filenames = [f"pdf_image{i+1}" for i in range(len(images))]
                
                # 上段の行（画像ファイル名行）
                filename_row = current_row
                # 下段の行（ファイルパスと画像行）
                image_row = current_row + 1
                
                # 行の高さを設定
                ws.row_dimensions[filename_row].height = 20  # ファイル名行は低め
                ws.row_dimensions[image_row].height = 75     # 画像行は高め
                
                # 下段にファイルパスを設定
                ws[f'A{image_row}'] = file_path
                
                # 各画像を処理
                for img_idx, image in enumerate(images):
                    try:
                        col_letter = chr(ord('B') + img_idx) if img_idx < 25 else f"A{chr(ord('A') + img_idx - 25)}"
                        
                        # 上段に画像ファイル名を設定
                        filename_cell = f'{col_letter}{filename_row}'
                        if img_idx < len(image_filenames):
                            ws[filename_cell] = image_filenames[img_idx]
                        else:
                            ws[filename_cell] = f"image{img_idx+1}"
                        
                        # 下段に画像を配置
                        image_cell = f'{col_letter}{image_row}'
                        
                        # 画像を100px×100pxにリサイズ
                        resized_image = resize_image_to_100px(image.copy())
                        
                        # 一時ファイルを作成（手動で削除）
                        temp_fd, temp_path = tempfile.mkstemp(suffix='.png')
                        temp_files.append(temp_path)  # 削除用にリストに追加
                        
                        try:
                            # ファイルディスクリプタを閉じる
                            os.close(temp_fd)
                            
                            # 画像を保存
                            resized_image.save(temp_path, 'PNG', optimize=True)
                            
                            # ファイルサイズをチェック（異常に大きい場合はスキップ）
                            if os.path.getsize(temp_path) > 10 * 1024 * 1024:  # 10MB制限
                                print(f"    → {img_idx+1}番目の画像: ファイルサイズが大きすぎます")
                                continue
                            
                            # Excelに画像を挿入
                            img = OpenpyxlImage(temp_path)
                            img.width = 100
                            img.height = 100
                            
                            ws.add_image(img, image_cell)
                            
                            filename = image_filenames[img_idx] if img_idx < len(image_filenames) else f"image{img_idx+1}"
                            print(f"  → {filename}: {col_letter}列に配置完了（ファイル名: {filename_row}行目、画像: {image_row}行目）")
                        except (OSError, PermissionError) as e:
                            print(f"  → {img_idx+1}番目の画像: ファイル操作エラー ({str(e)})")
                        except Exception as e:
                            print(f"  → {img_idx+1}番目の画像: エラー ({str(e)})")
                    except Exception as e:
                        print(f"  → {img_idx+1}番目の画像: 画像処理エラー ({str(e)})")
                
                print(f"  → 合計 {len(images)} 個の画像を処理（{filename_row}-{image_row}行目）")
                
                # 次のファイル用に行を2つ進める
                current_row += 2
                
            except Exception as e:
                print(f"  → ファイル処理エラー: {str(e)}")
                # エラーが発生してもカウンターは進める
                current_row += 2
                continue
        
        # Excelファイルを保存
        try:
            wb.save(output_path)
            print(f"Excelファイルを保存しました: {output_path}")
        except (PermissionError, OSError) as e:
            print(f"Excelファイル保存エラー: {str(e)}")
            # 代替ファイル名で保存を試行
            import time
            alt_filename = f"画像含有ファイル検索結果_{int(time.time())}.xlsx"
            alt_path = os.path.join(output_dir, alt_filename)
            wb.save(alt_path)
            print(f"代替ファイル名で保存しました: {alt_path}")
    
    finally:
        # 一時ファイルをクリーンアップ
        for temp_path in temp_files:
            try:
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
            except Exception:
                pass  # 削除できなくても続行

def save_to_csv_with_image_info(file_list, output_dir="result", output_filename="検索結果.csv"):
    """
    ファイルリストと画像情報をCSVファイルに保存する（実際のファイル名を使用）
    
    Args:
        file_list (list): ファイルパスのリスト
        output_dir (str): 出力ディレクトリ
        output_filename (str): 出力するCSVファイル名
    """
    if not file_list:
        print("保存するファイルがありません。")
        return
    
    # 出力ディレクトリを作成
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, output_filename)
    
    # CSVデータを準備
    csv_data = []
    
    print("CSV用データを作成中...")
    for file_path in file_list:
        try:
            print(f"画像情報抽出中: {file_path}")
            
            row = {'ファイルパス': file_path}
            
            if file_path.endswith('.docx'):
                media_filenames = get_media_filenames(file_path)
                # 全ての画像ファイル名を記録
                for i, filename in enumerate(media_filenames):
                    row[f'画像{i+1}_ファイル名'] = filename
                    
            elif file_path.endswith('.pdf'):
                images = extract_images_from_pdf(file_path)
                # PDFの場合は便宜的な名前を使用
                for i in range(len(images)):
                    row[f'画像{i+1}_ファイル名'] = f"pdf_image{i+1}"
            
            csv_data.append(row)
            
        except Exception as e:
            print(f"  → ファイル処理エラー: {str(e)}")
            # エラーの場合もファイルパスは記録
            csv_data.append({'ファイルパス': file_path})
            continue
    
    # DataFrameを作成してCSV保存
    try:
        df = pd.DataFrame(csv_data)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"CSVファイルを保存しました: {output_path}")
    except Exception as e:
        print(f"CSVファイル保存エラー: {str(e)}")

def main():
    # 検索対象のディレクトリを指定（現在のディレクトリから開始）
    search_directory = "."
    
    print(f"ディレクトリをクロール中: {os.path.abspath(search_directory)}")
    print("-" * 50)
    
    # ファイルを抽出
    files = extract_docx_pdf_files(search_directory)
    
    # 結果を表示
    if files:
        print(f"見つかったファイル数: {len(files)}")
        
        # 画像が含まれているファイルのみを抽出
        print("\n" + "-" * 50)
        print("画像が含まれているファイルをチェック中...")
        print("-" * 50)
        
        files_with_images = filter_files_with_images(files)
        
        print("\n" + "-" * 50)
        print(f"画像が含まれているファイル数: {len(files_with_images)}")
        print("\n画像が含まれているファイル:")
        for i, file_path in enumerate(files_with_images, 1):
            print(f"{i:3d}. {file_path}")
        
        # Excelファイルに保存（実際の画像ファイル名でヘッダー作成）
        if files_with_images:
            print("\n" + "-" * 50)
            print("画像ファイル情報を含むExcelファイルを作成中...")
            print("-" * 50)
            save_to_excel_with_images(files_with_images, "result", "検索結果.xlsx")
            
            print("\n" + "-" * 50)
            print("実際の画像ファイル名を使用したCSVファイルを作成中...")
            print("-" * 50)
            save_to_csv_with_image_info(files_with_images, "result", "検索結果.csv")
        else:
            print("\n画像が含まれているファイルがありませんでした。")
    else:
        print("該当するファイルが見つかりませんでした。")
    
    print("-" * 50)
    print("抽出完了")

if __name__ == "__main__":
    main() 